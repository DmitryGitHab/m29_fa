from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
import os
import uuid
import re

from starlette.requests import Request
from starlette.responses import HTMLResponse

from pydantic import BaseModel
from typing import List

app = FastAPI()

# Папка для загрузки файлов
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Словарь для хранения путей к загруженным файлам
uploaded_files = {}

# Монтируем статические файлы (CSS, JS)
app.mount("/static", StaticFiles(directory="frontend"), name="static")

# Главная страница
@app.get("/", response_class=HTMLResponse)
async def read_root():
    with open("frontend/index.html", "r", encoding="utf-8") as file:
        return HTMLResponse(content=file.read())

# Загрузка файла М-29
@app.post("/upload/m29")
async def upload_m29(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")

    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())

    uploaded_files["m29"] = file_path
    return {"message": "Файл М-29 успешно загружен", "file_path": file_path}

# Проверка состояния загрузки файлов
@app.get("/status")
async def get_status():
    return {
        "m29_loaded": "m29" in uploaded_files,
        "m29_path": uploaded_files.get("m29", ""),
        "ks2_loaded": "ks2" in uploaded_files,
        "ks2_path": uploaded_files.get("ks2", ""),
        "sap_loaded": "sap" in uploaded_files,
        "sap_path": uploaded_files.get("sap", ""),
    }



@app.get("/get_sheets")
async def get_sheets():
    if "m29" not in uploaded_files:
        raise HTTPException(status_code=400, detail="Файл М-29 не загружен")

    # Открываем файл и получаем список листов
    wb = openpyxl.load_workbook(uploaded_files["m29"], read_only=True)
    sheets = wb.sheetnames  # Получаем список имен листов

    return {"sheets": sheets}

# Модель для валидации входных данных
class UnwrapRequest(BaseModel):
    m29_name: str
    mtr_mask: List[str]

# Распаковка М-29
@app.post("/m_unwrap")
async def m_unwrap(request: UnwrapRequest):
    if "m29" not in uploaded_files:
        raise HTTPException(status_code=400, detail="Файл М-29 не загружен")

    m29_path = uploaded_files["m29"]
    wb = openpyxl.load_workbook(m29_path, data_only=True)
    sheet = wb[request.m29_name]  # Используем request.m29_name
    row = sheet.max_row
    column = sheet.max_column
    dest = {}

    for mm in request.mtr_mask:  # Используем request.mtr_mask
        for c in range(column):
            for r in range(row):
                val = sheet.cell(row=r + 1, column=c + 1).value
                if '.' not in str(val):
                    if mm in str(val):
                        val = re.split(",| |№", str(val))[-1]
                        if request.mtr_mask.index(mm) == 0:
                            for r2 in range(r + 1, row):
                                val2 = sheet.cell(row=r2 + 1, column=c + 1).value
                                if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                    count = round(sheet.cell(row=r2 + 1, column=c + 2).value, 4)
                                    break
                            if val in dest:
                                dest[val] = round(dest[val] + count, 4)
                            else:
                                dest[val] = count
                        else:
                            if request.mtr_mask[request.mtr_mask.index(mm)-1] not in str(val):
                                for r2 in range(r + 1, row):
                                    val2 = sheet.cell(row=r2 + 1, column=c + 1).value
                                    if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                        count = round(sheet.cell(row=r2 + 1, column=c + 2).value, 4)
                                        break
                                if val in dest:
                                    dest[val] = round(dest[val] + count, 4)
                                else:
                                    dest[val] = count

    wb0 = openpyxl.Workbook()
    Sheet_name = wb0.active

    Sheet_name.column_dimensions["A"].width = 25
    Sheet_name.column_dimensions["B"].width = 20

    Sheet_name["A1"] = "Номенклатурный номер"
    Sheet_name["B1"] = "Количество"

    for row, (key, hourly) in enumerate(dest.items(), start=2):
        Sheet_name[f"A{row}"] = key
        Sheet_name[f"B{row}"] = hourly

    result_path = os.path.join(UPLOAD_DIR, "расскрытая м29.xlsx")
    wb0.save(result_path)

    return FileResponse(result_path, filename="расскрытая м29.xlsx")




@app.get("/get_sheets_ks2")
async def get_sheets_ks2():
    if "ks2" not in uploaded_files:
        raise HTTPException(status_code=400, detail="Файл КС-2 не загружен")

    wb = openpyxl.load_workbook(uploaded_files["ks2"], read_only=True)
    sheets = wb.sheetnames
    print(f"Листы в файле КС-2: {sheets}")  # Логируем список листов
    return {"sheets": sheets}

@app.post("/upload/ks2")
async def upload_ks2(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")

    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())

    uploaded_files["ks2"] = file_path
    print(f"Файл КС-2 загружен: {file_path}")  # Логируем загрузку
    return {"message": "Файл КС-2 успешно загружен", "file_path": file_path}


@app.post("/compare_m29_ks2")
async def compare_m29_ks2(request: Request):
    try:
        body = await request.json()
        print("Тело запроса:", body)  # Логируем тело запроса

        if "m29_name" not in body or "ks2_name" not in body or "mtr_mask" not in body or "added_int" not in body:
            raise HTTPException(status_code=422, detail="Неверный формат данных")

        m29_name = body["m29_name"]
        ks2_name = body["ks2_name"]
        mtr_mask = body["mtr_mask"]
        added_int = body["added_int"]  # Получаем added_int из запроса

        if "m29" not in uploaded_files:
            raise HTTPException(status_code=400, detail="Файл М-29 не загружен")
        if "ks2" not in uploaded_files:
            raise HTTPException(status_code=400, detail="Файл КС-2 не загружен")

        # Обработка файла М-29
        wb_m = openpyxl.load_workbook(uploaded_files["m29"], data_only=True)
        sheet_m = wb_m[m29_name]
        row_m = sheet_m.max_row
        column_m = sheet_m.max_column
        dest_m = {}

        for mm in mtr_mask:
            for c in range(column_m):
                for r in range(row_m):
                    val = sheet_m.cell(row=r + 1, column=c + 1).value
                    if '.' not in str(val):
                        if mm in str(val):
                            val = re.split(",| |№", str(val))[-1]
                            if mtr_mask.index(mm) == 0:
                                for r2 in range(r + 1, row_m):
                                    val2 = sheet_m.cell(row=r2 + 1, column=c + 1).value
                                    if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                        count = round(sheet_m.cell(row=r2 + 1, column=c + 2).value, 4)
                                        break
                                if val in dest_m:
                                    dest_m[val] = round(dest_m[val] + count, 4)
                                else:
                                    dest_m[val] = count
                            else:
                                if mtr_mask[mtr_mask.index(mm)-1] not in str(val):
                                    for r2 in range(r + 1, row_m):
                                        val2 = sheet_m.cell(row=r2 + 1, column=c + 1).value
                                        if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                            count = round(sheet_m.cell(row=r2 + 1, column=c + 2).value, 4)
                                            break
                                    if val in dest_m:
                                        dest_m[val] = round(dest_m[val] + count, 4)
                                    else:
                                        dest_m[val] = count

        # Обработка файла КС-2
        wb_ks = openpyxl.load_workbook(uploaded_files["ks2"], data_only=True)
        sheet_ks = wb_ks[ks2_name]
        row_ks = sheet_ks.max_row
        column_ks = sheet_ks.max_column
        dest_ks = {}

        for mm in mtr_mask:
            for c in range(column_ks):
                for r in range(row_ks):
                    val = sheet_ks.cell(row=r + 1, column=c + 1).value
                    val2 = sheet_ks.cell(row=r + 1, column=c + 2 + added_int).value  # Используем added_int
                    if str(mm) in str(val):
                        val = re.split(",| |№", str(val))[-1]
                        if mtr_mask.index(mm) == 0:
                            if val in dest_ks:
                                dest_ks[val] = round(float(dest_ks[val]) + float(val2), 4)
                            else:
                                dest_ks[val] = round(float(val2), 4)
                        else:
                            if mtr_mask[mtr_mask.index(mm) - 1] not in str(val):
                                if val in dest_ks:
                                    dest_ks[val] = round(float(dest_ks[val]) + float(val2), 4)
                                else:
                                    dest_ks[val] = round(float(val2), 4)

        # Сравнение данных
        wrong_dict = {}
        for e in dest_m.keys():
            if e in dest_ks:
                if dest_m[e] != dest_ks[e]:
                    wrong_dict[e] = [dest_m[e], dest_ks[e]]
            else:
                wrong_dict[e] = [dest_m[e], 0]

        for e in dest_ks.keys():
            if e in dest_m:
                if dest_ks[e] != dest_m[e]:
                    if e not in wrong_dict:
                        wrong_dict[e] = [dest_m[e], dest_ks[e]]
            else:
                wrong_dict[e] = [0, dest_ks[e]]

        # Создание отчета
        wb0 = openpyxl.Workbook()
        Sheet_name = wb0.active

        Sheet_name.column_dimensions["A"].width = 25
        Sheet_name.column_dimensions["B"].width = 20
        Sheet_name.column_dimensions["C"].width = 20

        Sheet_name["A1"] = "Номенклатурный номер"
        Sheet_name["B1"] = "Количество М29"
        Sheet_name["C1"] = "Количество КС-2"

        for row, (key, hourly) in enumerate(wrong_dict.items(), start=2):
            Sheet_name[f"A{row}"] = key
            Sheet_name[f"B{row}"] = hourly[0]
            Sheet_name[f"C{row}"] = hourly[1]

        result_path = os.path.join(UPLOAD_DIR, "расхождения м29 и кс2.xlsx")
        wb0.save(result_path)

        return FileResponse(result_path, filename="расхождения м29 и кс2.xlsx")
    except Exception as e:
        print(f"Ошибка: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.post("/upload/sap")
async def upload_sap(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    file_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file.filename}")

    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())

    uploaded_files["sap"] = file_path
    print(f"Файл SAP загружен: {file_path}")  # Логируем загрузку
    return {"message": "Файл SAP успешно загружен", "file_path": file_path}


@app.post("/compare_m29_sap")
async def compare_m29_sap(request: Request):
    try:
        body = await request.json()
        print("Тело запроса:", body)  # Логируем тело запроса

        if "m29_name" not in body or "mtr_mask" not in body:
            raise HTTPException(status_code=422, detail="Неверный формат данных")

        m29_name = body["m29_name"]
        mtr_mask = body["mtr_mask"]

        if "m29" not in uploaded_files:
            raise HTTPException(status_code=400, detail="Файл М-29 не загружен")
        if "sap" not in uploaded_files:
            raise HTTPException(status_code=400, detail="Файл SAP не загружен")

        m29_path = uploaded_files["m29"]
        sap_path = uploaded_files["sap"]

        # Обработка файла М-29
        wb_m = openpyxl.load_workbook(m29_path, data_only=True)
        sheet_m = wb_m[m29_name]
        row_m = sheet_m.max_row
        column_m = sheet_m.max_column
        dest_m = {}

        for mm in mtr_mask:
            for c in range(column_m):
                for r in range(row_m):
                    val = sheet_m.cell(row=r + 1, column=c + 1).value
                    if '.' not in str(val):
                        if mm in str(val):
                            val = re.split(",| |№", str(val))[-1]
                            if mtr_mask.index(mm) == 0:
                                for r2 in range(r + 1, row_m):
                                    val2 = sheet_m.cell(row=r2 + 1, column=c + 1).value
                                    if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                        count = round(sheet_m.cell(row=r2 + 1, column=c + 2).value, 4)
                                        break
                                if val in dest_m:
                                    dest_m[val] = round(dest_m[val] + count, 4)
                                else:
                                    dest_m[val] = count
                            else:
                                if mtr_mask[mtr_mask.index(mm)-1] not in str(val):
                                    for r2 in range(r + 1, row_m):
                                        val2 = sheet_m.cell(row=r2 + 1, column=c + 1).value
                                        if ('Х' or 'x' or 'Х' or 'х') in str(val2):
                                            count = round(sheet_m.cell(row=r2 + 1, column=c + 2).value, 4)
                                            break
                                    if val in dest_m:
                                        dest_m[val] = round(dest_m[val] + count, 4)
                                    else:
                                        dest_m[val] = count

        # Обработка файла SAP
        wb_sap = openpyxl.load_workbook(sap_path, data_only=True)
        sheet_sap = wb_sap.active
        row_sap = sheet_sap.max_row
        column_sap = sheet_sap.max_column
        dest_sap = {}

        for c in range(column_sap):
            for r in range(row_sap):
                mtr = sheet_sap.cell(row=r + 1, column=c + 1).value
                if "Материал" == str(mtr):
                    mtr_c_n = c + 1
                if "Кол-во" == str(mtr):
                    mtr_c_c = c + 1

        for mm in mtr_mask:
            for r in range(row_sap):
                val = sheet_sap.cell(row=r + 1, column=mtr_c_n).value
                val2 = sheet_sap.cell(row=r + 1, column=mtr_c_c).value
                if mm in str(val):
                    val = re.split(",| |№", str(val))[-1]
                    if mtr_mask.index(mm) == 0:
                        if val in dest_sap:
                            dest_sap[val] = round(float(dest_sap[val]) + float(val2), 4)
                        else:
                            dest_sap[val] = float(val2)
                    else:
                        if mtr_mask[mtr_mask.index(mm) - 1] not in str(val):
                            if val in dest_sap:
                                dest_sap[val] = round(float(dest_sap[val]) + float(val2), 4)
                            else:
                                dest_sap[val] = float(val2)

        # Сравнение данных
        wrong_dict = {}
        for e in dest_m.keys():
            if e in dest_sap:
                if dest_m[e] != dest_sap[e]:
                    wrong_dict[e] = [dest_m[e], dest_sap[e]]
            else:
                wrong_dict[e] = [dest_m[e], 0]

        for e in dest_sap.keys():
            if e in dest_m:
                if dest_sap[e] != dest_m[e]:
                    if e not in wrong_dict:
                        wrong_dict[e] = [dest_m[e], dest_sap[e]]
            else:
                wrong_dict[e] = [0, dest_sap[e]]

        # Создание отчета
        wb0 = openpyxl.Workbook()
        Sheet_name = wb0.active

        Sheet_name.column_dimensions["A"].width = 25
        Sheet_name.column_dimensions["B"].width = 20
        Sheet_name.column_dimensions["C"].width = 20

        Sheet_name["A1"] = "Номенклатурный номер"
        Sheet_name["B1"] = "Количество М29"
        Sheet_name["C1"] = "Количество SAP"

        for row, (key, hourly) in enumerate(wrong_dict.items(), start=2):
            Sheet_name[f"A{row}"] = key
            Sheet_name[f"B{row}"] = hourly[0]
            Sheet_name[f"C{row}"] = hourly[1]

        result_path = os.path.join(UPLOAD_DIR, "расхождения м29 и sap.xlsx")
        wb0.save(result_path)

        return FileResponse(result_path, filename="расхождения м29 и sap.xlsx")
    except Exception as e:
        print(f"Ошибка: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")
